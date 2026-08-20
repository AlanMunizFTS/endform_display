from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory

import cv2
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PilImage, ImageDraw

from paths_config import REPORTS_DIR


STATUS_COLUMNS = ("OK", "NOK", "FOK", "FNOK", "Total")
TRACEABILITY_STATUS_COLUMNS = ("OK", "NOK", "Total")
TRACEABILITY_REPORT_KIND = "ok_nok_by_jsn"
DEFAULT_HISTORIC_REPORT_DEFECT_CLASS = "wrinkle"
DEFAULT_HISTORIC_REPORT_ANGLE = "side"


def parse_jsn_datetime(jsn):
    text = str(jsn or "").strip()
    if len(text) < 17 or not text[5:17].isdigit():
        raise ValueError("JSN does not contain MMDDYYHHMMSS at positions 6-17")

    date_token = text[5:11]
    time_token = text[11:17]
    month = int(date_token[0:2])
    day = int(date_token[2:4])
    year = 2000 + int(date_token[4:6])
    hour = int(time_token[0:2])
    minute = int(time_token[2:4])
    second = int(time_token[4:6])
    return datetime(year, month, day, hour, minute, second)


def _normalize_traceability_result(value):
    result = str(value or "").strip().upper()
    if result in {"OK", "FOK"}:
        return "OK"
    if result in {"NOK", "FNOK"}:
        return "NOK"
    return None


def _fetch_traceability_piece_rows(db_client):
    return list(
        db_client.fetch(
            "SELECT jsn, model_result FROM piece_result ORDER BY jsn"
        )
        or []
    )


def build_ok_nok_traceability_report(db_client):
    rows = _fetch_traceability_piece_rows(db_client)
    warnings = []
    valid_entries = []

    for row in rows:
        jsn = row.get("jsn") if hasattr(row, "get") else row[0]
        model_result = row.get("model_result") if hasattr(row, "get") else row[1]
        normalized_result = _normalize_traceability_result(model_result)
        if normalized_result is None:
            warnings.append(
                {
                    "jsn": jsn,
                    "reason": f"Unsupported model_result: {model_result}",
                }
            )
            continue

        try:
            captured_at = parse_jsn_datetime(jsn)
        except ValueError as exc:
            warnings.append({"jsn": jsn, "reason": str(exc)})
            continue

        valid_entries.append(
            {
                "jsn": str(jsn),
                "result": normalized_result,
                "captured_at": captured_at,
            }
        )

    day_map = {}
    hour_map = {}
    total_ok = 0
    total_nok = 0

    for entry in valid_entries:
        captured_at = entry["captured_at"]
        result = entry["result"]
        if result == "OK":
            total_ok += 1
        else:
            total_nok += 1

        day_key = captured_at.date()
        day_counts = day_map.setdefault(day_key, {"OK": 0, "NOK": 0})
        day_counts[result] += 1

        hour_key = captured_at.replace(minute=0, second=0, microsecond=0)
        hour_counts = hour_map.setdefault(hour_key, {"OK": 0, "NOK": 0})
        hour_counts[result] += 1

    start_at = min((entry["captured_at"] for entry in valid_entries), default=None)
    end_at = max((entry["captured_at"] for entry in valid_entries), default=None)

    day_rows = []
    for day_key in sorted(day_map):
        counts = day_map[day_key]
        total = counts["OK"] + counts["NOK"]
        day_rows.append(
            {
                "date": day_key,
                "OK": counts["OK"],
                "NOK": counts["NOK"],
                "Total": total,
                "pct_ok": counts["OK"] / total if total else 0,
                "pct_nok": counts["NOK"] / total if total else 0,
            }
        )

    hour_rows = []
    if start_at is not None and end_at is not None:
        current_hour = start_at.replace(minute=0, second=0, microsecond=0)
        final_hour = end_at.replace(minute=0, second=0, microsecond=0)
        while current_hour <= final_hour:
            counts = hour_map.get(current_hour, {"OK": 0, "NOK": 0})
            total = counts["OK"] + counts["NOK"]
            hour_rows.append(
                {
                    "date": current_hour.date(),
                    "hour": current_hour.hour,
                    "range": f"{current_hour:%H}:00 - {current_hour:%H}:59",
                    "OK": counts["OK"],
                    "NOK": counts["NOK"],
                    "Total": total,
                    "pct_ok": counts["OK"] / total if total else 0,
                    "pct_nok": counts["NOK"] / total if total else 0,
                }
            )
            current_hour += timedelta(hours=1)

    if not rows:
        warnings.append({"jsn": "", "reason": "No piece_result rows available"})

    total = total_ok + total_nok
    return {
        "kind": TRACEABILITY_REPORT_KIND,
        "start_at": start_at,
        "end_at": end_at,
        "summary": {
            "OK": total_ok,
            "NOK": total_nok,
            "Total": total,
            "pct_ok": total_ok / total if total else 0,
            "pct_nok": total_nok / total if total else 0,
        },
        "day_rows": day_rows,
        "hour_rows": hour_rows,
        "warnings": warnings,
    }


def _format_report_filename(start_at, end_at):
    if start_at is None or end_at is None:
        raise ValueError("Stats report requires a valid DB date range")

    return (
        f"reporte_{start_at:%Y%m%d}_{end_at:%Y%m%d}_{start_at:%H%M}_{end_at:%H%M}.xlsx"
    )


def _format_traceability_filename(created_at=None):
    timestamp = created_at or datetime.now()
    return f"desglose_ok_nok_{timestamp:%Y%m%d_%H%M%S}.xlsx"


def _format_historic_image_report_filename(created_at=None):
    timestamp = created_at or datetime.now()
    return f"reporte_imagenes_historico_{timestamp:%Y%m%d_%H%M%S}.xlsx"


def _build_historic_image_report_image_header(endform_type, class_name):
    base_parts = [
        str(endform_type or "").strip(),
        str(class_name or "").strip(),
    ]
    base_label = "-".join(part for part in base_parts if part)
    return base_label or "Image"


def _apply_headers(sheet, headers, header_fill, header_font, center):
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center


def _write_count_row(
    sheet,
    row_idx,
    values,
    total_fill,
    total_font,
    center,
    percent_columns,
):
    for col_idx, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = center
        if row_idx == 2 and sheet.title == "Resumen OK-NOK":
            cell.fill = total_fill
            cell.font = total_font
        if col_idx in percent_columns:
            cell.number_format = "0.00%"


def _build_traceability_workbook(db_client):
    report_data = build_ok_nok_traceability_report(db_client)
    workbook = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="404040")
    total_fill = PatternFill(fill_type="solid", fgColor="6E6E6E")
    header_font = Font(color="FFFFFF", bold=True)
    total_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen OK-NOK"
    summary_headers = [
        "Periodo inicio",
        "Periodo fin",
        "OK",
        "NOK",
        "Total",
        "% OK",
        "% NOK",
    ]
    _apply_headers(summary_sheet, summary_headers, header_fill, header_font, center)
    summary = report_data["summary"]
    _write_count_row(
        summary_sheet,
        2,
        [
            report_data["start_at"].strftime("%Y-%m-%d %H:%M:%S")
            if report_data["start_at"]
            else "",
            report_data["end_at"].strftime("%Y-%m-%d %H:%M:%S")
            if report_data["end_at"]
            else "",
            summary["OK"],
            summary["NOK"],
            summary["Total"],
            summary["pct_ok"],
            summary["pct_nok"],
        ],
        total_fill,
        total_font,
        center,
        percent_columns={6, 7},
    )

    day_sheet = workbook.create_sheet("Por dia")
    count_headers = ["Fecha", *TRACEABILITY_STATUS_COLUMNS, "% OK", "% NOK"]
    _apply_headers(day_sheet, count_headers, header_fill, header_font, center)
    for row_idx, row in enumerate(report_data["day_rows"], start=2):
        _write_count_row(
            day_sheet,
            row_idx,
            [
                row["date"].strftime("%Y-%m-%d"),
                row["OK"],
                row["NOK"],
                row["Total"],
                row["pct_ok"],
                row["pct_nok"],
            ],
            total_fill,
            total_font,
            center,
            percent_columns={5, 6},
        )

    hour_sheet = workbook.create_sheet("Por hora")
    hour_headers = ["Fecha", "Hora", "Rango", *TRACEABILITY_STATUS_COLUMNS, "% OK", "% NOK"]
    _apply_headers(hour_sheet, hour_headers, header_fill, header_font, center)
    for row_idx, row in enumerate(report_data["hour_rows"], start=2):
        _write_count_row(
            hour_sheet,
            row_idx,
            [
                row["date"].strftime("%Y-%m-%d"),
                f"{row['hour']:02d}",
                row["range"],
                row["OK"],
                row["NOK"],
                row["Total"],
                row["pct_ok"],
                row["pct_nok"],
            ],
            total_fill,
            total_font,
            center,
            percent_columns={7, 8},
        )

    warning_sheet = workbook.create_sheet("Advertencias")
    _apply_headers(warning_sheet, ["JSN", "Motivo"], header_fill, header_font, center)
    for row_idx, warning in enumerate(report_data["warnings"], start=2):
        jsn_cell = warning_sheet.cell(row=row_idx, column=1, value=warning.get("jsn", ""))
        reason_cell = warning_sheet.cell(row=row_idx, column=2, value=warning.get("reason", ""))
        jsn_cell.alignment = left
        reason_cell.alignment = left

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_width = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )
            sheet.column_dimensions[column_letter].width = min(max(max_width + 2, 12), 42)

    return workbook


def _add_stats_matrix_chart(sheet, matrix_rows):
    data_row_count = 0
    for row in matrix_rows:
        if row.get("is_total"):
            break
        data_row_count += 1

    if data_row_count < 1:
        return

    max_row = data_row_count + 1
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Stats Matrix"
    chart.y_axis.title = "Pieces"
    chart.x_axis.title = "Class Name"
    chart.height = 7.5
    chart.width = 16

    data = Reference(sheet, min_col=2, max_col=5, min_row=1, max_row=max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "H2")


def _add_stats_matrix_sheet(workbook, matrix_rows, sheet=None):
    sheet = sheet or workbook.create_sheet("Stats")
    sheet.title = "Stats"
    header_fill = PatternFill(fill_type="solid", fgColor="404040")
    total_fill = PatternFill(fill_type="solid", fgColor="6E6E6E")
    header_font = Font(color="FFFFFF", bold=True)
    total_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Class Name", *STATUS_COLUMNS]
    _apply_headers(sheet, headers, header_fill, header_font, center)

    for row_idx, row in enumerate(matrix_rows, start=2):
        values = [
            row.get("class_name", ""),
            row.get("OK", 0),
            row.get("NOK", 0),
            row.get("FOK", 0),
            row.get("FNOK", 0),
            row.get("Total", 0),
        ]
        is_total = bool(row.get("is_total"))
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left if col_idx == 1 else center
            if is_total:
                cell.fill = total_fill
                cell.font = total_font

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 28
    for column_letter in ("B", "C", "D", "E", "F"):
        sheet.column_dimensions[column_letter].width = 12
    _add_stats_matrix_chart(sheet, matrix_rows)
    return sheet


def export_ok_nok_traceability_report(db_client, output_dir=None, created_at=None):
    report_dir = Path(output_dir or REPORTS_DIR)
    report_dir.mkdir(parents=True, exist_ok=True)
    workbook = _build_traceability_workbook(db_client)
    filename = _format_traceability_filename(created_at=created_at)
    output_path = report_dir / filename
    workbook.save(output_path)
    return str(output_path)


def export_combined_traceability_report(controller, db_client=None, output_dir=None, created_at=None):
    db = db_client or getattr(controller.display, "db", None)
    if db is None:
        raise ValueError("No database connection available")

    matrix_data = controller.build_piece_stats_report(db_client=db)
    matrix_rows = list(matrix_data.get("rows") or [])
    if not matrix_rows:
        raise ValueError("No piece stats available to export")

    report_dir = Path(output_dir or REPORTS_DIR)
    report_dir.mkdir(parents=True, exist_ok=True)
    workbook = _build_traceability_workbook(db)
    _add_stats_matrix_sheet(workbook, matrix_rows)
    filename = _format_traceability_filename(created_at=created_at)
    output_path = report_dir / filename
    workbook.save(output_path)
    return str(output_path)


def export_stats_report(controller, db_client=None, output_dir=None):
    report_data = controller.build_piece_stats_report(db_client=db_client)
    matrix_rows = list(report_data.get("rows") or [])
    if not matrix_rows:
        raise ValueError("No piece stats available to export")

    report_dir = Path(output_dir or REPORTS_DIR)
    report_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    _add_stats_matrix_sheet(workbook, matrix_rows, sheet=workbook.active)

    filename = _format_report_filename(
        report_data.get("start_at"),
        report_data.get("end_at"),
    )
    output_path = report_dir / filename
    workbook.save(output_path)
    return str(output_path)


def _make_piece_contact_sheet(
    image_paths,
    output_path,
    tile_size=110,
    padding=6,
    label_text="",
):
    cols = 4
    rows = 2
    label_height = 30 if label_text else 0
    width = cols * tile_size + (cols - 1) * padding
    height = label_height + rows * tile_size + (rows - 1) * padding
    canvas = PilImage.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(canvas)

    if label_text:
        draw.rectangle((0, 0, width, label_height), fill=(31, 78, 121))
        draw.text((8, 8), label_text, fill="white")

    for idx in range(cols * rows):
        col = idx % cols
        row = idx // cols
        x = col * (tile_size + padding)
        y = label_height + row * (tile_size + padding)
        box = (x, y, x + tile_size, y + tile_size)

        if idx >= len(image_paths):
            continue

        try:
            with PilImage.open(image_paths[idx]) as source:
                image = source.convert("RGB")
                image.thumbnail((tile_size, tile_size), PilImage.LANCZOS)
                paste_x = x + (tile_size - image.width) // 2
                paste_y = y + (tile_size - image.height) // 2
                canvas.paste(image, (paste_x, paste_y))
        except Exception:
            draw.rectangle(box, fill=(245, 245, 245), outline=(190, 190, 190))
            draw.line((x + 12, y + 12, x + tile_size - 12, y + tile_size - 12), fill=(170, 170, 170), width=3)
            draw.line((x + tile_size - 12, y + 12, x + 12, y + tile_size - 12), fill=(170, 170, 170), width=3)

        draw.rectangle(box, outline=(210, 210, 210), width=1)

    canvas.save(output_path)


def _set_image_anchor(workbook_image, row_idx, col_idx, width, height, margin_px=6):
    marker = AnchorMarker(
        col=col_idx - 1,
        row=row_idx - 1,
        colOff=pixels_to_EMU(margin_px),
        rowOff=pixels_to_EMU(margin_px),
    )
    workbook_image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height)),
    )


def _load_historic_report_overlays(controller, image_names, chunk_size=500):
    overlay_getter = getattr(controller, "get_model_overlays_for_images", None)
    if not callable(overlay_getter):
        return {}

    names = [str(name or "").strip() for name in image_names if str(name or "").strip()]
    overlays_by_image = {}
    resolved_chunk_size = max(1, int(chunk_size or 500))
    for start_idx in range(0, len(names), resolved_chunk_size):
        chunk = names[start_idx : start_idx + resolved_chunk_size]
        try:
            chunk_overlays = overlay_getter(chunk)
        except Exception:
            continue
        if isinstance(chunk_overlays, dict):
            overlays_by_image.update(chunk_overlays)
    return overlays_by_image


def _is_historic_report_angle_image(img_name, angle):
    filename = Path(str(img_name or "")).name.lower()
    normalized_angle = str(angle or "").strip().lower()
    return bool(normalized_angle) and normalized_angle in filename


def _filter_historic_report_overlays(overlays_by_image, defect_class, angle):
    normalized_class = str(defect_class or "").strip().lower()
    filtered = {}
    for img_name, overlays in (overlays_by_image or {}).items():
        if not _is_historic_report_angle_image(img_name, angle):
            continue
        matching_overlays = [
            overlay
            for overlay in (overlays or [])
            if str(overlay.get("class_name") or "").strip().lower()
            == normalized_class
        ]
        if matching_overlays:
            filtered[img_name] = matching_overlays
    return filtered


def _prepare_historic_report_image(
    controller,
    img_name,
    historic_dir,
    annotated_dir,
    overlays,
    temp_dir,
    image_idx,
    target_size=110,
    allow_annotated_fallback=True,
):
    historic_path = Path(historic_dir) / img_name
    annotated_path = Path(annotated_dir) / img_name if annotated_dir else None
    overlay_renderer = getattr(
        getattr(controller, "display", None),
        "_draw_model_overlays",
        None,
    )

    if overlays and historic_path.exists() and callable(overlay_renderer):
        try:
            base_image = cv2.imread(str(historic_path), cv2.IMREAD_COLOR)
            if base_image is not None:
                source_h, source_w = base_image.shape[:2]
                target_size = max(1, int(target_size or 110))
                interpolation = (
                    cv2.INTER_AREA
                    if source_h > target_size or source_w > target_size
                    else cv2.INTER_LINEAR
                )
                base_image = cv2.resize(
                    base_image,
                    (target_size, target_size),
                    interpolation=interpolation,
                )
                marked_image = overlay_renderer(
                    base_image.copy(),
                    overlays,
                    source_w,
                    source_h,
                )
                marked_path = Path(temp_dir) / f"marked_{image_idx:08d}.png"
                if marked_image is not None and cv2.imwrite(str(marked_path), marked_image):
                    return marked_path
        except Exception:
            pass

    if (
        allow_annotated_fallback
        and annotated_path is not None
        and annotated_path.exists()
    ):
        return annotated_path
    return historic_path


def _extract_historic_batch_jsn(batch):
    for img_name in batch or []:
        filename = Path(str(img_name or "")).name
        if filename:
            return filename.split("_", 1)[0]
    return ""


def _load_historic_filtered_verdicts(
    controller,
    historic_index,
    defect_class,
    angle,
    chunk_size=500,
):
    db_client = getattr(getattr(controller, "display", None), "db", None)
    if db_client is None or not callable(getattr(db_client, "fetch", None)):
        return {}

    normalized_class = str(defect_class or "").strip().lower()
    normalized_angle = str(angle or "").strip().lower()
    if not normalized_class or not normalized_angle:
        return {}

    filtered_images_by_jsn = {}
    jsn_by_image = {}
    for batch in historic_index or []:
        jsn = _extract_historic_batch_jsn(batch)
        if not jsn:
            continue
        filtered_images = {
            str(img_name)
            for img_name in (batch or [])
            if _is_historic_report_angle_image(img_name, normalized_angle)
        }
        if not filtered_images:
            continue
        filtered_images_by_jsn[jsn] = filtered_images
        for img_name in filtered_images:
            jsn_by_image[img_name] = jsn

    filtered_image_names = list(jsn_by_image)
    queried_images = set()
    nok_jsns = set()
    resolved_chunk_size = max(1, int(chunk_size or 500))
    for start_idx in range(0, len(filtered_image_names), resolved_chunk_size):
        chunk = filtered_image_names[start_idx : start_idx + resolved_chunk_size]
        try:
            rows = db_client.fetch(
                "SELECT img_name, class_name FROM model_results "
                "WHERE img_name = ANY(%s) AND LOWER(TRIM(class_name)) = %s",
                (chunk, normalized_class),
            )
        except Exception:
            continue
        if not isinstance(rows, (list, tuple)):
            continue
        queried_images.update(chunk)
        for row in rows:
            if not hasattr(row, "get"):
                continue
            img_name = str(row.get("img_name") or "").strip()
            class_name = str(row.get("class_name") or "").strip().lower()
            if class_name != normalized_class:
                continue
            jsn = jsn_by_image.get(img_name)
            if jsn:
                nok_jsns.add(jsn)

    verdicts = {}
    for jsn, filtered_images in filtered_images_by_jsn.items():
        if not filtered_images.issubset(queried_images):
            continue
        verdicts[jsn] = "NOK" if jsn in nok_jsns else "OK"
    return verdicts


def export_historic_image_table_report(
    controller,
    output_dir=None,
    historic_dir=None,
    annotated_dir=None,
    created_at=None,
    endform_type="",
    class_name="",
    defect_class=DEFAULT_HISTORIC_REPORT_DEFECT_CLASS,
    angle=DEFAULT_HISTORIC_REPORT_ANGLE,
    pieces_per_group=4,
    images_per_piece=7,
    progress_callback=None,
):
    historic_index = list(reversed(controller._load_historic_index(force_rescan=True) or []))
    if not historic_index:
        raise ValueError("No historic images available to export")

    source_dir = Path(historic_dir or controller._get_export_historic_dir())
    if annotated_dir is None:
        annotated_dir_getter = getattr(controller, "_get_annotated_historic_dir", None)
        candidate_annotated_dir = (
            annotated_dir_getter() if callable(annotated_dir_getter) else None
        )
        if isinstance(candidate_annotated_dir, (str, Path)):
            annotated_dir = candidate_annotated_dir
    annotated_source_dir = Path(annotated_dir) if annotated_dir else None
    report_dir = Path(output_dir or REPORTS_DIR)
    report_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Piezas"
    sheet.sheet_view.showGridLines = False
    verdict_sheet = workbook.create_sheet("Piezas con veredicto")
    verdict_sheet.sheet_view.showGridLines = False

    pieces_per_group = max(1, int(pieces_per_group or 4))
    images_per_piece = max(1, int(images_per_piece or 7))
    defect_class = (
        str(defect_class or DEFAULT_HISTORIC_REPORT_DEFECT_CLASS).strip().lower()
    )
    angle = str(angle or DEFAULT_HISTORIC_REPORT_ANGLE).strip().lower()
    tile_size = 110
    padding = 6
    image_margin_px = 6
    composite_width = 4 * tile_size + 3 * padding
    composite_height = 2 * tile_size + padding

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    good_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    bad_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    good_font = Font(color="006100", bold=True)
    bad_font = Font(color="9C0006", bold=True)

    total_pieces = len(historic_index)
    total_groups = max(1, (total_pieces + pieces_per_group - 1) // pieces_per_group)
    total_cols = pieces_per_group + 1
    verdict_total_cols = pieces_per_group * 2 + 1
    first_data_row = 4
    image_header = _build_historic_image_report_image_header(
        endform_type,
        defect_class,
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = sheet.cell(row=1, column=1, value="PART-BY-PART RESULT SPLIT")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, total_cols + 1):
        title_range_cell = sheet.cell(row=1, column=col_idx)
        title_range_cell.fill = title_fill
        title_range_cell.border = table_border
    sheet.row_dimensions[1].height = 24

    group_header = sheet.cell(row=2, column=1, value="Pieza agrupada")
    group_header.fill = header_fill
    group_header.font = header_font
    group_header.alignment = center
    group_header.border = table_border

    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols)
    image_header_cell = sheet.cell(row=2, column=2, value=image_header)
    image_header_cell.fill = header_fill
    image_header_cell.font = header_font
    image_header_cell.alignment = center
    for col_idx in range(2, total_cols + 1):
        cell = sheet.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = table_border

    column_headers = ["Grupo"] + ["Capturas"] * pieces_per_group
    for col_idx, header in enumerate(column_headers, start=1):
        cell = sheet.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = table_border

    sheet.column_dimensions["A"].width = 24
    for col_idx in range(2, total_cols + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = (
            composite_width + image_margin_px * 2
        ) / 7.0
    sheet.freeze_panes = "B4"

    for group_idx in range(total_groups):
        row_idx = first_data_row + group_idx
        sheet.row_dimensions[row_idx].height = (composite_height + image_margin_px * 2) * 0.75
        for col_idx in range(1, total_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = table_border
        sheet.cell(
            row=row_idx,
            column=1,
            value=f"Pieza agrupada #{group_idx + 1}",
        )

    verdict_sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=verdict_total_cols,
    )
    verdict_title_cell = verdict_sheet.cell(
        row=1,
        column=1,
        value=f"PART-BY-PART {defect_class.upper()} {angle.upper()} VERDICT",
    )
    verdict_title_cell.fill = title_fill
    verdict_title_cell.font = title_font
    verdict_title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, verdict_total_cols + 1):
        cell = verdict_sheet.cell(row=1, column=col_idx)
        cell.fill = title_fill
        cell.border = table_border
    verdict_sheet.row_dimensions[1].height = 24

    verdict_group_header = verdict_sheet.cell(
        row=2,
        column=1,
        value="Pieza agrupada",
    )
    verdict_group_header.fill = header_fill
    verdict_group_header.font = header_font
    verdict_group_header.alignment = center
    verdict_group_header.border = table_border

    verdict_sheet.merge_cells(
        start_row=2,
        start_column=2,
        end_row=2,
        end_column=verdict_total_cols,
    )
    verdict_image_header = verdict_sheet.cell(row=2, column=2, value=image_header)
    verdict_image_header.fill = header_fill
    verdict_image_header.font = header_font
    verdict_image_header.alignment = center
    for col_idx in range(2, verdict_total_cols + 1):
        cell = verdict_sheet.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = table_border

    verdict_headers = ["Grupo"]
    for _position in range(pieces_per_group):
        verdict_headers.extend(("Veredicto", "Capturas"))
    for col_idx, header in enumerate(verdict_headers, start=1):
        cell = verdict_sheet.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = table_border

    verdict_sheet.column_dimensions["A"].width = 24
    for position in range(pieces_per_group):
        verdict_col = position * 2 + 2
        image_col = verdict_col + 1
        verdict_sheet.column_dimensions[get_column_letter(verdict_col)].width = 12
        verdict_sheet.column_dimensions[get_column_letter(image_col)].width = (
            composite_width + image_margin_px * 2
        ) / 7.0
    verdict_sheet.freeze_panes = "B4"

    for group_idx in range(total_groups):
        row_idx = first_data_row + group_idx
        verdict_sheet.row_dimensions[row_idx].height = (
            composite_height + image_margin_px * 2
        ) * 0.75
        group_cell = verdict_sheet.cell(
            row=row_idx,
            column=1,
            value=f"Pieza agrupada #{group_idx + 1}",
        )
        group_cell.alignment = center
        group_cell.border = table_border
        for col_idx in range(2, verdict_total_cols + 1):
            cell = verdict_sheet.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = table_border

    if progress_callback:
        progress_callback(0, total_pieces, "Preparing image report")

    selected_image_names = [
        img_name
        for batch in historic_index
        for img_name in list(batch or [])[:images_per_piece]
    ]
    overlays_by_image = _filter_historic_report_overlays(
        _load_historic_report_overlays(
            controller,
            selected_image_names,
        ),
        defect_class,
        angle,
    )
    model_verdicts = _load_historic_filtered_verdicts(
        controller,
        historic_index,
        defect_class,
        angle,
    )

    with TemporaryDirectory() as temp_dir:
        prepared_image_idx = 0
        for piece_idx, batch in enumerate(historic_index):
            group_idx = piece_idx // pieces_per_group
            position_in_group = piece_idx % pieces_per_group
            data_row = first_data_row + group_idx
            piece_col = position_in_group + 2
            verdict_col = position_in_group * 2 + 2
            verdict_piece_col = verdict_col + 1

            selected_images = list(batch or [])[:images_per_piece]
            image_paths = []
            for img_name in selected_images:
                prepared_image_idx += 1
                image_paths.append(
                    _prepare_historic_report_image(
                        controller,
                        img_name,
                        source_dir,
                        annotated_source_dir,
                        overlays_by_image.get(img_name) or [],
                        temp_dir,
                        prepared_image_idx,
                        target_size=tile_size,
                        allow_annotated_fallback=False,
                    )
                )
            composite_path = Path(temp_dir) / f"piece_{piece_idx + 1:06d}.png"
            _make_piece_contact_sheet(
                image_paths,
                composite_path,
                tile_size=tile_size,
                padding=padding,
            )

            workbook_image = WorkbookImage(str(composite_path))
            workbook_image.width = composite_width
            workbook_image.height = composite_height
            _set_image_anchor(
                workbook_image,
                data_row,
                piece_col,
                composite_width,
                composite_height,
                margin_px=image_margin_px,
            )
            sheet.add_image(workbook_image)

            verdict_workbook_image = WorkbookImage(str(composite_path))
            verdict_workbook_image.width = composite_width
            verdict_workbook_image.height = composite_height
            _set_image_anchor(
                verdict_workbook_image,
                data_row,
                verdict_piece_col,
                composite_width,
                composite_height,
                margin_px=image_margin_px,
            )
            verdict_sheet.add_image(verdict_workbook_image)

            jsn = _extract_historic_batch_jsn(batch)
            verdict = model_verdicts.get(jsn)
            verdict_cell = verdict_sheet.cell(
                row=data_row,
                column=verdict_col,
                value=verdict,
            )
            verdict_cell.alignment = center
            verdict_cell.border = table_border
            if verdict == "OK":
                verdict_cell.fill = good_fill
                verdict_cell.font = good_font
            elif verdict == "NOK":
                verdict_cell.fill = bad_fill
                verdict_cell.font = bad_font

            if progress_callback:
                progress_callback(piece_idx + 1, total_pieces, "Adding pieces to workbook")

        filename = _format_historic_image_report_filename(created_at=created_at)
        output_path = report_dir / filename
        workbook.save(output_path)

    return str(output_path)
