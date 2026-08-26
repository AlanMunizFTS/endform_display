import datetime
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook
from PIL import Image as PilImage

from report_exporter import (
    export_combined_traceability_report,
    export_historic_image_table_report,
    export_ok_nok_traceability_report,
    export_stats_report,
    parse_jsn_datetime,
)


class FakeTraceabilityDB:
    def __init__(self, rows):
        self.rows = rows

    def fetch(self, query, data=None):
        normalized = " ".join(query.split())
        if normalized == "SELECT jsn, model_result FROM piece_result ORDER BY jsn":
            return sorted(self.rows, key=lambda row: row["jsn"])
        raise AssertionError(f"Unhandled query: {normalized}")


class TestReportExporter(unittest.TestCase):
    def assert_stats_matrix_chart(self, sheet, expected_last_data_row):
        self.assertEqual(len(sheet._charts), 1)
        chart = sheet._charts[0]
        self.assertEqual(chart.type, "col")
        self.assertEqual(chart.grouping, "stacked")
        self.assertEqual(chart.overlap, 100.0)
        self.assertEqual(chart.anchor._from.col, 7)
        self.assertEqual(chart.anchor._from.row, 1)

        expected_category_range = f"'Stats'!$A$2:$A${expected_last_data_row}"
        expected_value_ranges = [
            f"'Stats'!${column}$2:${column}${expected_last_data_row}"
            for column in ("B", "C", "D", "E")
        ]
        self.assertEqual(len(chart.series), len(expected_value_ranges))
        for series, expected_value_range in zip(chart.series, expected_value_ranges):
            self.assertEqual(series.val.numRef.f, expected_value_range)
            self.assertEqual(series.cat.numRef.f, expected_category_range)

    def test_parse_jsn_datetime_reads_date_and_time_tokens(self):
        parsed = parse_jsn_datetime("218620514260607413863")

        self.assertEqual(parsed, datetime.datetime(2026, 5, 14, 6, 7, 41))

    def test_export_ok_nok_traceability_report_groups_by_day_and_hour(self):
        db = FakeTraceabilityDB(
            [
                {"jsn": "218620514260607413863", "model_result": "OK"},
                {"jsn": "218620514260645413864", "model_result": "OK"},
                {"jsn": "218620514260807413865", "model_result": "NOK"},
                {"jsn": "bad-jsn", "model_result": "NOK"},
            ]
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = export_ok_nok_traceability_report(
                db,
                tmp_dir,
                created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
            )

            path = Path(output_path)
            self.assertEqual(path.name, "desglose_ok_nok_20260514_093000.xlsx")

            workbook = load_workbook(path)
            self.assertIn("Resumen OK-NOK", workbook.sheetnames)
            self.assertIn("Por dia", workbook.sheetnames)
            self.assertIn("Por hora", workbook.sheetnames)
            self.assertIn("Advertencias", workbook.sheetnames)

            summary = workbook["Resumen OK-NOK"]
            self.assertEqual(summary["C2"].value, 2)
            self.assertEqual(summary["D2"].value, 1)
            self.assertEqual(summary["E2"].value, 3)

            by_day = workbook["Por dia"]
            self.assertEqual(by_day["A2"].value, "2026-05-14")
            self.assertEqual(by_day["B2"].value, 2)
            self.assertEqual(by_day["C2"].value, 1)

            by_hour = workbook["Por hora"]
            self.assertEqual(by_hour["B2"].value, "06")
            self.assertEqual(by_hour["D2"].value, 2)
            self.assertEqual(by_hour["B3"].value, "07")
            self.assertEqual(by_hour["F3"].value, 0)
            self.assertEqual(by_hour["B4"].value, "08")
            self.assertEqual(by_hour["E4"].value, 1)

            warnings = workbook["Advertencias"]
            self.assertEqual(warnings["A2"].value, "bad-jsn")

    def test_export_stats_report_creates_workbook_with_expected_name_and_table(self):
        controller = MagicMock()
        controller.build_piece_stats_report.return_value = {
            "rows": [
                {"class_name": "OK", "OK": 15, "NOK": 0, "FOK": 0, "FNOK": 0, "Total": 15},
                {"class_name": "split", "OK": 0, "NOK": 5, "FOK": 0, "FNOK": 3, "Total": 8},
                {"class_name": "Total", "OK": 15, "NOK": 5, "FOK": 0, "FNOK": 3, "Total": 23, "is_total": True},
            ],
            "start_at": datetime.datetime(2026, 3, 25, 9, 0),
            "end_at": datetime.datetime(2026, 3, 25, 12, 0),
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = export_stats_report(controller, output_dir=tmp_dir)

            path = Path(output_path)
            self.assertTrue(path.exists())
            self.assertEqual(
                path.name,
                "reporte_20260325_20260325_0900_1200.xlsx",
            )

            workbook = load_workbook(path)
            sheet = workbook["Stats"]
            self.assertEqual(sheet["A1"].value, "Class Name")
            self.assertEqual(sheet["B1"].value, "OK")
            self.assertEqual(sheet["F1"].value, "Total")
            self.assertEqual(sheet["A2"].value, "OK")
            self.assertEqual(sheet["B2"].value, 15)
            self.assertEqual(sheet["A4"].value, "Total")
            self.assertEqual(sheet["F4"].value, 23)
            self.assert_stats_matrix_chart(sheet, expected_last_data_row=3)

    def test_export_combined_traceability_report_includes_stats_matrix_sheet(self):
        db = FakeTraceabilityDB(
            [
                {"jsn": "218620514260607413863", "model_result": "OK"},
                {"jsn": "218620514260645413864", "model_result": "OK"},
                {"jsn": "218620514260807413865", "model_result": "NOK"},
            ]
        )
        controller = MagicMock()
        controller.display.db = db
        controller.build_piece_stats_report.return_value = {
            "rows": [
                {"class_name": "OK", "OK": 2, "NOK": 0, "FOK": 0, "FNOK": 0, "Total": 2},
                {"class_name": "scratch", "OK": 0, "NOK": 1, "FOK": 0, "FNOK": 0, "Total": 1},
                {"class_name": "Total", "OK": 2, "NOK": 1, "FOK": 0, "FNOK": 0, "Total": 3, "is_total": True},
            ],
            "start_at": datetime.datetime(2026, 5, 14, 6, 0),
            "end_at": datetime.datetime(2026, 5, 14, 8, 0),
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = export_combined_traceability_report(
                controller,
                db_client=db,
                output_dir=tmp_dir,
                created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
            )

            path = Path(output_path)
            self.assertEqual(path.name, "desglose_ok_nok_20260514_093000.xlsx")
            workbook = load_workbook(path)
            self.assertEqual(
                workbook.sheetnames,
                ["Resumen OK-NOK", "Por dia", "Por hora", "Advertencias", "Stats"],
            )
            self.assertEqual(workbook["Resumen OK-NOK"]["E2"].value, 3)
            stats_sheet = workbook["Stats"]
            self.assertEqual(stats_sheet["F4"].value, 3)
            self.assert_stats_matrix_chart(stats_sheet, expected_last_data_row=3)

    def test_export_stats_report_skips_chart_when_matrix_has_only_total_row(self):
        controller = MagicMock()
        controller.build_piece_stats_report.return_value = {
            "rows": [
                {
                    "class_name": "Total",
                    "OK": 0,
                    "NOK": 0,
                    "FOK": 0,
                    "FNOK": 0,
                    "Total": 0,
                    "is_total": True,
                },
            ],
            "start_at": datetime.datetime(2026, 3, 25, 9, 0),
            "end_at": datetime.datetime(2026, 3, 25, 12, 0),
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = export_stats_report(controller, output_dir=tmp_dir)

            workbook = load_workbook(output_path)
            sheet = workbook["Stats"]
            self.assertEqual(sheet["A2"].value, "Total")
            self.assertEqual(sheet["F2"].value, 0)
            self.assertEqual(sheet._charts, [])

    def test_export_stats_report_requires_valid_db_date_range(self):
        controller = MagicMock()
        controller.build_piece_stats_report.return_value = {
            "rows": [
                {"class_name": "OK", "OK": 1, "NOK": 0, "FOK": 0, "FNOK": 0, "Total": 1},
                {"class_name": "Total", "OK": 1, "NOK": 0, "FOK": 0, "FNOK": 0, "Total": 1, "is_total": True},
            ],
            "start_at": None,
            "end_at": None,
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            with self.assertRaisesRegex(ValueError, "valid DB date range"):
                export_stats_report(controller, output_dir=tmp_dir)

    def test_export_historic_image_table_report_groups_four_pieces_per_row(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            historic_dir = Path(tmp_dir) / "historic"
            report_dir = Path(tmp_dir) / "reports"
            historic_dir.mkdir()

            historic_index = []
            for piece_idx in range(8):
                jsn = f"11861022070165{piece_idx:03d}"
                batch = []
                for cam_idx in range(7):
                    image_name = f"{jsn}_Cam{cam_idx + 1}_Side_OK.png"
                    image_path = historic_dir / image_name
                    PilImage.new(
                        "RGB",
                        (32, 32),
                        (piece_idx * 30, cam_idx * 30, 120),
                    ).save(image_path)
                    batch.append(image_name)
                historic_index.append(batch)

            controller = MagicMock()
            controller._load_historic_index.return_value = historic_index
            controller._get_export_historic_dir.return_value = str(historic_dir)
            controller.display.db.fetch.return_value = [
                {
                    "img_name": (
                        f"11861022070165{piece_idx:03d}_Cam1_Side_OK.png"
                    ),
                    "class_name": "wrinkle",
                }
                for piece_idx in range(8)
                if piece_idx % 2 == 1
            ] + [
                {
                    "img_name": "11861022070165000_Cam1_Side_OK.png",
                    "class_name": "dent",
                },
                {
                    "img_name": "11861022070165000_Cam2_Diag_OK.png",
                    "class_name": "wrinkle",
                },
            ]
            captured_contact_sheets = []

            def fake_make_contact_sheet(image_paths, output_path, **kwargs):
                captured_contact_sheets.append(
                    {
                        "images": [Path(image_path).name for image_path in image_paths],
                        "label_text": kwargs.get("label_text"),
                    }
                )
                PilImage.new("RGB", (32, 32), (120, 120, 120)).save(output_path)

            with patch(
                "report_exporter._make_piece_contact_sheet",
                side_effect=fake_make_contact_sheet,
            ):
                output_path = export_historic_image_table_report(
                    controller,
                    output_dir=report_dir,
                    created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
                    endform_type="mush",
                    class_name="split",
                )

            path = Path(output_path)
            self.assertEqual(
                path.name,
                "historic_image_report_20260514_093000.xlsx",
            )
            workbook = load_workbook(path)
            sheet = workbook["Pieces"]
            self.assertEqual(sheet["A1"].value, "PART-BY-PART RESULT SPLIT")
            self.assertEqual(sheet["A2"].value, "Grouped piece")
            self.assertEqual(sheet["B2"].value, "mush-wrinkle")
            self.assertEqual(sheet["A3"].value, "Group")
            self.assertEqual(sheet["B3"].value, "Captures")
            self.assertEqual(sheet["C3"].value, "Captures")
            self.assertEqual(sheet["D3"].value, "Captures")
            self.assertEqual(sheet["E3"].value, "Captures")
            self.assertEqual(sheet["A4"].value, "Grouped piece #1")
            self.assertEqual(sheet["A5"].value, "Grouped piece #2")
            self.assertEqual(sheet["A6"].value, None)
            self.assertEqual(len(sheet._images), 8)
            anchors = [
                (image.anchor._from.row, image.anchor._from.col)
                for image in sheet._images
            ]
            self.assertEqual(
                anchors,
                [
                    (3, 1),
                    (3, 2),
                    (3, 3),
                    (3, 4),
                    (4, 1),
                    (4, 2),
                    (4, 3),
                    (4, 4),
                ],
            )
            self.assertEqual(
                [item["label_text"] for item in captured_contact_sheets],
                [None] * 8,
            )
            self.assertTrue(
                all(len(item["images"]) == 7 for item in captured_contact_sheets)
            )
            self.assertEqual(
                [
                    image_name.split("_")[1]
                    for image_name in captured_contact_sheets[0]["images"]
                ],
                [f"Cam{cam_idx}" for cam_idx in range(1, 8)],
            )

            verdict_sheet = workbook["Pieces with verdict"]
            self.assertEqual(
                verdict_sheet["A1"].value,
                "PART-BY-PART WRINKLE SIDE VERDICT",
            )
            self.assertEqual(verdict_sheet["A4"].value, "Grouped piece #1")
            self.assertEqual(verdict_sheet["A5"].value, "Grouped piece #2")
            self.assertEqual(verdict_sheet["B3"].value, "Verdict")
            self.assertEqual(verdict_sheet["C3"].value, "Captures")
            self.assertEqual(
                [
                    verdict_sheet.cell(row=4, column=col_idx).value
                    for col_idx in (2, 4, 6, 8)
                ],
                ["NOK", "OK", "NOK", "OK"],
            )
            self.assertEqual(
                [
                    verdict_sheet.cell(row=5, column=col_idx).value
                    for col_idx in (2, 4, 6, 8)
                ],
                ["NOK", "OK", "NOK", "OK"],
            )
            self.assertEqual(
                [
                    (image.anchor._from.row, image.anchor._from.col)
                    for image in verdict_sheet._images
                ],
                [
                    (3, 2),
                    (3, 4),
                    (3, 6),
                    (3, 8),
                    (4, 2),
                    (4, 4),
                    (4, 6),
                    (4, 8),
                ],
            )
            verdict_query = controller.display.db.fetch.call_args[0][0]
            self.assertIn("FROM model_results", verdict_query)
            self.assertIn("LOWER(TRIM(class_name))", verdict_query)

            jsn_sheet = workbook["JSN by round"]
            self.assertEqual(
                [jsn_sheet.cell(row=1, column=col_idx).value for col_idx in range(1, 7)],
                [
                    "Grouped piece",
                    "Round 1",
                    "Round 2",
                    "Round 3",
                    "Round 4",
                    "Original OK/NOK value",
                ],
            )
            self.assertEqual(jsn_sheet["A2"].value, "Grouped piece 1")
            self.assertEqual(jsn_sheet["A3"].value, "Grouped piece 2")
            self.assertEqual(
                [jsn_sheet.cell(row=2, column=col_idx).value for col_idx in range(2, 6)],
                [f"11861022070165{piece_idx:03d}" for piece_idx in range(7, 3, -1)],
            )
            self.assertEqual(
                [jsn_sheet.cell(row=3, column=col_idx).value for col_idx in range(2, 6)],
                [f"11861022070165{piece_idx:03d}" for piece_idx in range(3, -1, -1)],
            )
            self.assertEqual(jsn_sheet["F2"].value, None)
            self.assertEqual(jsn_sheet["F3"].value, None)
            self.assertEqual(len(jsn_sheet._images), 0)

    def test_export_historic_image_table_report_uses_defect_markings(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            historic_dir = Path(tmp_dir) / "historic"
            annotated_dir = Path(tmp_dir) / "annotated"
            report_dir = Path(tmp_dir) / "reports"
            historic_dir.mkdir()
            annotated_dir.mkdir()

            jsn = "118610220701650001"
            overlay_image = f"{jsn}_Cam1_Side_OK.png"
            annotated_image = f"{jsn}_Cam2_Diag_NOK.png"
            PilImage.new("RGB", (32, 32), (255, 255, 255)).save(
                historic_dir / overlay_image
            )
            PilImage.new("RGB", (32, 32), (255, 255, 255)).save(
                historic_dir / annotated_image
            )
            PilImage.new("RGB", (32, 32), (0, 0, 255)).save(
                annotated_dir / annotated_image
            )

            controller = MagicMock()
            controller._load_historic_index.return_value = [
                [overlay_image, annotated_image]
            ]
            controller._get_export_historic_dir.return_value = str(historic_dir)
            controller._get_annotated_historic_dir.return_value = str(annotated_dir)
            controller.get_model_overlays_for_images.return_value = {
                overlay_image: [
                    {
                        "geometry_type": "bbox",
                        "coordinates": [1, 1, 8, 8],
                        "class_name": "wrinkle",
                    },
                    {
                        "geometry_type": "bbox",
                        "coordinates": [2, 2, 6, 6],
                        "class_name": "dent",
                    },
                ],
                annotated_image: [
                    {
                        "geometry_type": "bbox",
                        "coordinates": [1, 1, 8, 8],
                        "class_name": "wrinkle",
                    }
                ],
            }
            controller.display.db.fetch.return_value = [
                {
                    "img_name": annotated_image,
                    "class_name": "wrinkle",
                }
            ]

            def fake_draw_overlays(image, overlays, source_w, source_h):
                self.assertEqual(len(overlays), 1)
                self.assertEqual(overlays[0]["class_name"], "wrinkle")
                self.assertEqual((source_w, source_h), (32, 32))
                image[:, :] = (0, 0, 255)
                return image

            controller.display._draw_model_overlays.side_effect = fake_draw_overlays
            captured_pixels = []
            captured_labels = []

            def fake_make_contact_sheet(image_paths, output_path, **kwargs):
                for image_path in image_paths:
                    with PilImage.open(image_path) as source:
                        captured_pixels.append(
                            source.convert("RGB").getpixel((0, 0))
                        )
                captured_labels.append(kwargs.get("label_text"))
                PilImage.new("RGB", (32, 32), (120, 120, 120)).save(output_path)

            with patch(
                "report_exporter._make_piece_contact_sheet",
                side_effect=fake_make_contact_sheet,
            ):
                side_output_path = export_historic_image_table_report(
                    controller,
                    output_dir=report_dir,
                    created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
                    endform_type="mush",
                    class_name="split",
                )

            self.assertEqual(captured_pixels, [(255, 0, 0), (255, 255, 255)])
            self.assertEqual(captured_labels, [None])
            controller.display._draw_model_overlays.assert_called_once()
            side_workbook = load_workbook(side_output_path)
            self.assertEqual(
                side_workbook["Pieces with verdict"]["B4"].value,
                "OK",
            )

            captured_pixels.clear()
            captured_labels.clear()
            controller.display._draw_model_overlays.reset_mock()
            controller.display._draw_model_overlays.side_effect = fake_draw_overlays

            with patch(
                "report_exporter._make_piece_contact_sheet",
                side_effect=fake_make_contact_sheet,
            ):
                diag_output_path = export_historic_image_table_report(
                    controller,
                    output_dir=report_dir,
                    created_at=datetime.datetime(2026, 5, 14, 9, 31, 0),
                    endform_type="mush",
                    defect_class="wrinkle",
                    angle="diag",
                )

            self.assertEqual(captured_pixels, [(255, 255, 255), (255, 0, 0)])
            self.assertEqual(captured_labels, [None])
            controller.display._draw_model_overlays.assert_called_once()
            diag_workbook = load_workbook(diag_output_path)
            self.assertEqual(
                diag_workbook["Pieces with verdict"]["B4"].value,
                "NOK",
            )

    def test_export_historic_image_table_report_keeps_incomplete_final_group(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            historic_dir = Path(tmp_dir) / "historic"
            report_dir = Path(tmp_dir) / "reports"
            historic_dir.mkdir()

            historic_index = []
            for piece_idx in range(5):
                jsn = f"11861022070165{piece_idx:03d}"
                image_name = f"{jsn}_Cam1_Side_OK.png"
                PilImage.new("RGB", (32, 32), (120, 120, 120)).save(
                    historic_dir / image_name
                )
                historic_index.append([image_name])

            controller = MagicMock()
            controller._load_historic_index.return_value = historic_index
            controller._get_export_historic_dir.return_value = str(historic_dir)
            controller.display.db.fetch.return_value = [
                {
                    "img_name": (
                        f"11861022070165{piece_idx:03d}_Cam1_Side_OK.png"
                    ),
                    "class_name": "wrinkle",
                }
                for piece_idx in range(5)
                if piece_idx % 2 == 1
            ]

            output_path = export_historic_image_table_report(
                controller,
                output_dir=report_dir,
                created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
                endform_type="mush",
                class_name="split",
            )

            workbook = load_workbook(output_path)
            sheet = workbook["Pieces"]
            self.assertEqual(sheet["A4"].value, "Grouped piece #1")
            self.assertEqual(sheet["A5"].value, "Grouped piece #2")
            self.assertEqual(sheet["C5"].value, None)
            self.assertEqual(sheet["D5"].value, None)
            self.assertEqual(sheet["E5"].value, None)
            self.assertEqual(
                [
                    (image.anchor._from.row, image.anchor._from.col)
                    for image in sheet._images
                ],
                [
                    (3, 1),
                    (3, 2),
                    (3, 3),
                    (3, 4),
                    (4, 1),
                ],
            )

            verdict_sheet = workbook["Pieces with verdict"]
            self.assertEqual(verdict_sheet["A4"].value, "Grouped piece #1")
            self.assertEqual(verdict_sheet["A5"].value, "Grouped piece #2")
            self.assertEqual(
                [
                    verdict_sheet.cell(row=4, column=col_idx).value
                    for col_idx in (2, 4, 6, 8)
                ],
                ["OK", "NOK", "OK", "NOK"],
            )
            self.assertEqual(verdict_sheet["B5"].value, "OK")
            self.assertEqual(verdict_sheet["D5"].value, None)
            self.assertEqual(verdict_sheet["F5"].value, None)
            self.assertEqual(verdict_sheet["H5"].value, None)

            jsn_sheet = workbook["JSN by round"]
            self.assertEqual(jsn_sheet["A3"].value, "Grouped piece 2")
            self.assertEqual(jsn_sheet["B3"].value, "11861022070165000")
            self.assertEqual(jsn_sheet["C3"].value, None)
            self.assertEqual(jsn_sheet["D3"].value, None)
            self.assertEqual(jsn_sheet["E3"].value, None)
            self.assertEqual(jsn_sheet["F3"].value, None)

    def test_export_historic_image_table_report_uses_display_piece_order(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            historic_dir = Path(tmp_dir) / "historic"
            report_dir = Path(tmp_dir) / "reports"
            historic_dir.mkdir()

            internal_index = []
            for jsn in (
                "118610220701658060005",
                "118610220701657170004",
                "118610220701656250003",
            ):
                image_name = f"{jsn}_Cam1_Side_OK.png"
                PilImage.new("RGB", (32, 32), (120, 120, 120)).save(
                    historic_dir / image_name
                )
                internal_index.append([image_name])

            controller = MagicMock()
            controller._load_historic_index.return_value = internal_index
            controller._get_export_historic_dir.return_value = str(historic_dir)
            captured_first_images = []

            def fake_make_contact_sheet(image_paths, output_path, **_kwargs):
                captured_first_images.append(Path(image_paths[0]).name)
                PilImage.new("RGB", (32, 32), (120, 120, 120)).save(output_path)

            with patch(
                "report_exporter._make_piece_contact_sheet",
                side_effect=fake_make_contact_sheet,
            ):
                export_historic_image_table_report(
                    controller,
                    output_dir=report_dir,
                    created_at=datetime.datetime(2026, 5, 14, 9, 30, 0),
                    endform_type="mush",
                    class_name="split",
                )

            self.assertEqual(
                captured_first_images[0],
                "118610220701656250003_Cam1_Side_OK.png",
            )


if __name__ == "__main__":
    unittest.main(verbosity=2)
